# import datetime
#
# user_input = input("enter goal\n")
# input_list = user_input.split(":")
#
# goal = input_list[0]
# deadline = input_list[1]
#
# deadline_date = datetime.datetime.strptime(deadline, "%d.%m.%y")
#
# today_date = datetime.datetime.today()
#
# time_till = deadline_date - today_date
#
# print(f"Dear User Time remaining for your goal : {goal} is {time_till.days} days")
#
import openpyxl


inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    print(supplier_name)

    # calculation for number of products
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1

print(products_per_supplier)
